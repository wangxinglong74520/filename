#!/usr/bin/env python
# -*- coding: utf-8 -*-
# @File    : test.py
# @Time    : 2021/2/28 18:30
# @Author  : Merle
# @Site    : 
"""

"""
import os
import time
import unittest

import openpyxl
from ddt import ddt, data, unpack
from selenium import webdriver
from HTMLTestRunner import HTMLTestRunner

# 读取Excel
def read_excel():
    work = openpyxl.load_workbook(os.getcwd()+'\\test.xlsx')
    sheet = work['login']
    for row in range(2, sheet.max_row + 1):  # 行数
        temp = []
        for col in range(1, sheet.max_column + 1):  # 列数
            temp.append(sheet.cell(row, col).value)
        print(temp)
    return temp


@ddt
class Test(unittest.TestCase):

    @data(*read_excel())
    # @unpack
    def test_jenkins_login(self,password):
        print(password)
        # driver = webdriver.Chrome()
        # driver.get('http://192.168.17.1:8888/jenkins/login')
        # driver.find_element_by_name('j_username').send_keys(username)
        # driver.find_element_by_name('j_password').send_keys(password)
        # time.sleep(1)
        # driver.find_element_by_name('Submit').click()


if __name__ == '__main__':
    unittest.main()
    # testcases = unittest.defaultTestLoader.discover(os.getcwd(),"*.py")
    # filename = open('log.html','wb')
    # runner = HTMLTestRunner(stream=filename,verbosity=2,
    #                    title='jenkins',description="报告详情")
    # runner.run(testcases)