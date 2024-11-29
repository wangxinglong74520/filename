import os
import re

import xlrd
import xlwt
from openpyxl import load_workbook


def rend_excel():
    file_path = 'TestCase_Export_20240219104615.xlsx'
    wb = xlrd.open_workbook(file_path)
    sheet_names = wb.sheet_names()
    ws = wb.sheet_by_name(sheet_names[0])
    # 获取所有的行
    nrows = ws.nrows
    # 获取所有的列
    result = []
    ncols = ws.ncols
    # 通过循环获取到表格中所有用例，id和名称
    # type_lists = ['01 电话', '02 聊天', '03 视频通话']
    # type_lists = ['01 电话', '02 聊天', '03 视频通话', '04 相机', '05 相册', '06 奖励', '08 应用市场', '09 秒表', '10 闹钟', '11 天气',
    #               '14 找手表', '15 手表挂失', '16 表盘', '18 积分系统（卡卡乐园）', '19 华为听书', '20 应用管理', '21 锁屏', '22 空间清理', '23 手电筒',
    #               '24 公告板', '26 应用权限管理', '28 三方应用', '29 应用专项', '31 个人中心', '32 短彩信',
    #               '33 端云交互', '34 等级系统', '35 主题','36 个性秀','37 空间', '43 摇一摇', '45 OEM定制', '46隐私中心&自动接听', '47畅联', '01 账号与绑定', '02 设置-新',
    #               '03 Launcher', '04 上课禁用', '05 防沉迷', '06 开关机', '07 按键', '08 超级省电（新UX框架）', '09 电源管理', '10 热保护',
    #               '11 演示版本']
    type_lists = ['01 电话', '02 聊天', '03 视频通话', '04 相机', '05 相册', '06 奖励', '08 应用市场', '09 秒表', '10 闹钟', '11 天气',
                  '14 找手表', '15 手表挂失', '16 表盘', '18 积分系统（卡卡乐园）', '19 华为听书', '20 应用管理', '21 锁屏', '22 数据清理', '23 手电筒',
                  '24 公告板', '26 应用权限管理', '28 三方应用', '29 应用专项', '31 个人中心', '32 短彩信',
                  '34 等级系统', '35 主题', '36 个性秀', '37 空间', '41 流量管理', '42 三方应用流量开关', '44 聊天装扮', '45 OEM定制',
                  '46 隐私中心&自动接听', '47 畅连', '48 语音助手', '01 账号与绑定', '02 设置-新',
                  '03 Launcher', '06 开关机', '10 热保护', '11 演示版本', '14 恢复出厂设置', '15 分享能力', '19 交互框架改版', '20 离线定位',
                  '27 电源管理', '28 夜间模式+语音push', '29 push功能', '35 青少年模式', '36  断网提醒', '37 特殊模式改版', '38健康使用手表', '04 上课禁用',
                  '39续航管理', '08省电模式', '40设备信号质量']
    type_li = []
    for i in range(nrows):
        case_type = ws.cell_value(i, 0)
        if case_type in type_lists and case_type != '':
            type_li.append((case_type, i))
    print(type_li)
    type_name = {}
    for i in range(len(type_li) - 1):
        # print(type_li[i][0], type_li[i][1], type_li[i + 1][1])
        for j in range(type_li[i][1], type_li[i + 1][1]):
            case_id = ws.cell_value(j, 2)
            case_name = ws.cell_value(j, 3)
            case_level = ws.cell_value(j, 4)
            if case_id != '':
                type_name.setdefault(type_li[i][0], []).append((case_id, case_name, case_level))
    return type_name


def compare_excel():
    type_name = rend_excel()
    file_path = r'case_nopass.xlsx'
    wb = xlrd.open_workbook(file_path)
    sheels = wb.sheet_names()
    ws = wb.sheet_by_name(sheels[-1])
    print(sheels[-1])
    nrows = ws.nrows
    num = 0
    type_name_data = {}
    # 执行人
    executor_tuple = {}
    no_find = []
    ok_find = []
    for i in range(1, nrows):
        cass_id = ws.cell_value(i, 0)
        result = ws.cell_value(i, 6)
        result_contents = ws.cell_value(i, 7)
        executor = ws.cell_value(i, 8)
        # aotutest_result = ws.cell_value(i, 3)
        # executor_tuple.setdefault(executor, []).append(aotutest_result)
        if cass_id != '':
            for i, y in type_name.items():
                for x in y:
                    if cass_id == x[0]:
                        ok_find.append(cass_id)
                        x = list(x)
                        x.append(result)
                        x.append(result_contents)
                        x.append(executor)
                        # print(x)
                        type_name_data.setdefault(str(i).split(' ')[-1], []).append(x)
                        num += 1
    print(type_name_data)
    for i in range(1, nrows):
        cass_id = ws.cell_value(i, 0)
        if cass_id not in ok_find:
            no_find.append(cass_id)

    print('本次全量用例执行报告用例数为：', num)
    print('本次全量用例：', nrows - 1)
    print('没有找到的用例', no_find)
    # 计算执行条数和自动化执行率
    # for x, y in executor_tuple.items():
    #     auto_all = len(y)
    #     auto_true = y.count('True')
    #     auto_false = y.count('False')
    #     print(f"{x}本轮自动化执行用例数：{auto_all},自动执行数{auto_true},非自动执行数{auto_false},自动化执行率：{auto_true / auto_all:.2%}")
    # print(executor_tuple)
    return type_name_data


def module_classify():
    """用例模块分类"""
    type_name = rend_excel()
    file_path = r'case_nopass.xlsx'
    wb = xlrd.open_workbook(file_path)
    sheels = wb.sheet_names()
    ws = wb.sheet_by_name(sheels[-1])
    print(sheels[-1])
    nrows = ws.nrows
    print(nrows, '---------------')
    num = 0
    n = 0
    type_name_data = {}
    # 执行人
    executor_tuple = {}
    for i in range(1, nrows):
        cass_id = ws.cell_value(i, 0)
        if cass_id != '':
            for i, y in type_name.items():
                for x in y:
                    if cass_id == x[0]:
                        x = list(x)
                        # print(x)
                        type_name_data.setdefault(str(i).split(' ')[-1], []).append(x)
                        num += 1
            n += 1
    print('本次全量用例执行报告用例数为：', num)
    print('本次全量用例数：', n)
    print('本次全量用例：', nrows - 1)
    # 计算执行条数和自动化执行率
    for x, y in executor_tuple.items():
        auto_all = len(y)
        auto_true = y.count('True')
        auto_false = y.count('False')
        print(f"{x}本轮自动化执行用例数：{auto_all},自动执行数{auto_true},非自动执行数{auto_false},自动化执行率：{auto_true / auto_all:.2%}")
    # print(executor_tuple)
    # print(type_name_data)
    return type_name_data


def write_excel_module_classify():
    type_name_data = module_classify()
    csv_file = '全量自动化用例测试报告.xls'
    if os.path.exists(csv_file):
        os.remove(csv_file)
    wb = xlwt.Workbook()
    sheet = wb.add_sheet('Sheet1', cell_overwrite_ok=True)
    # 设置字体
    style = xlwt.XFStyle()
    # 字体
    font = xlwt.Font()
    font.name = '微软雅黑'
    font.height = 20 * 12
    style.font = font
    # 边框
    border = xlwt.Borders()
    border.top = 1
    border.bottom = 1
    border.left = 1
    border.right = 1
    style.borders = border
    # 对齐方式
    alignment1 = xlwt.Alignment()
    alignment1.vert = 0x01
    alignment1.horz = 0x01
    style.alignment = alignment1

    # 设置格式2
    style2 = xlwt.XFStyle()
    # 字体
    font = xlwt.Font()
    font.name = '微软雅黑'
    font.height = 20 * 12
    style2.font = font
    # 边框
    border = xlwt.Borders()
    border.top = 1
    border.bottom = 1
    border.left = 1
    border.right = 1
    style2.borders = border
    # 对齐方式
    alignment1 = xlwt.Alignment()
    alignment1.vert = 0x01
    alignment1.horz = 0x02
    style2.alignment = alignment1

    # 设置格式2
    style3 = xlwt.XFStyle()
    # 字体
    font = xlwt.Font()
    font.name = '微软雅黑'
    font.height = 20 * 12
    style3.font = font
    # 边框
    border = xlwt.Borders()
    border.top = 1
    border.bottom = 1
    border.left = 1
    border.right = 1
    style3.borders = border
    # 对齐方式
    alignment1 = xlwt.Alignment()
    alignment1.vert = 0x01
    alignment1.horz = 0x02
    style3.alignment = alignment1
    # 设置背景颜色
    pattern = xlwt.Pattern()
    pattern.pattern = xlwt.Pattern.SOLID_PATTERN
    pattern.pattern_fore_colour = 3
    style3.pattern = pattern

    # 设置格式4
    style4 = xlwt.XFStyle()
    # 字体
    font = xlwt.Font()
    font.name = '微软雅黑'
    font.height = 20 * 12
    style4.font = font
    # 边框
    border = xlwt.Borders()
    border.top = 1
    border.bottom = 1
    border.left = 1
    border.right = 1
    style4.borders = border
    # 对齐方式
    alignment1 = xlwt.Alignment()
    alignment1.vert = 0x01
    alignment1.horz = 0x02
    style4.alignment = alignment1
    # 设置背景颜色
    pattern = xlwt.Pattern()
    pattern.pattern = xlwt.Pattern.SOLID_PATTERN
    pattern.pattern_fore_colour = 2
    style4.pattern = pattern
    # 获取的用例字典去重
    module_name_deduplicate = {}
    module_name_deduplicate_list = []
    for n, m in type_name_data.items():
        for d in m:
            if d not in module_name_deduplicate_list:
                module_name_deduplicate_list.append(d)
                module_name_deduplicate.setdefault(n, []).append(d)

    num_pass = 0
    num_fail = 0
    lie_list = []
    module_names = module_name_deduplicate.keys()
    for i in module_names:
        nm = len(module_name_deduplicate.get(i))
        lie_list.append(nm)
    print(len(lie_list), '-------------------')
    print(lie_list)
    n = 1
    h = 0
    # 将所有发现的问题写道一起
    for i, y in module_name_deduplicate.items():
        """
        合并单元格使用ws.merge_cells()方法，参数介绍：
        start_row: 开始合并的行。
        start_column: 开始合并的行。
        end_row: 结束合并的列。
        end_column: 结束合并的列。这四个值所在的行/列都会被合并（闭区间）。"""
        # sheet.write(n, 0, i, style2)
        if i == '超级省电（新UX框架）':
            i = '超级省电'
        elif i == '积分系统（卡卡乐园）':
            i = '卡卡乐园'
        elif i == '超级省电（新UX框架）':
            i = '超级省电'
        elif i == '设置-新':
            i = '设置'
        sheet.write_merge(n, n + lie_list[h] - 1, 0, 0, i, style2)
        print(n, n + lie_list[h] - 1, 0, 0, i)
        sheet.write(0, 0, '模块', style)
        sheet.write(0, 1, '用例编号', style)
        sheet.write(n, 2, '用例名称', style)
        sheet.write(n, 3, '用例等级', style2)
        for x in y:
            sheet.write(n, 1, x[0], style)
            # write_merge合并单元格
            # sheet.write_merge(n, n, 1, 2, x[0], style)
            sheet.write(n, 2, x[1], style)
            # sheet.write_merge(n, n, 3, 4, x[1], style)
            sheet.write(n, 3, x[2], style2)
            n += 1
        h += 1
    wb.save(csv_file)


def write_excel():
    type_name_data = compare_excel()
    csv_file = '全量自动化用例测试报告.xls'
    if os.path.exists(csv_file):
        os.remove(csv_file)
    wb = xlwt.Workbook()
    sheet = wb.add_sheet('Sheet1', cell_overwrite_ok=True)
    # 设置字体
    style = xlwt.XFStyle()
    # 字体
    font = xlwt.Font()
    font.name = '微软雅黑'
    font.height = 20 * 12
    style.font = font
    # 边框
    border = xlwt.Borders()
    border.top = 1
    border.bottom = 1
    border.left = 1
    border.right = 1
    style.borders = border
    # 对齐方式
    alignment1 = xlwt.Alignment()
    alignment1.vert = 0x01
    alignment1.horz = 0x01
    style.alignment = alignment1

    # 设置格式2
    style2 = xlwt.XFStyle()
    # 字体
    font = xlwt.Font()
    font.name = '微软雅黑'
    font.height = 20 * 12
    style2.font = font
    # 边框
    border = xlwt.Borders()
    border.top = 1
    border.bottom = 1
    border.left = 1
    border.right = 1
    style2.borders = border
    # 对齐方式
    alignment1 = xlwt.Alignment()
    alignment1.vert = 0x01
    alignment1.horz = 0x02
    style2.alignment = alignment1

    # 设置格式2
    style3 = xlwt.XFStyle()
    # 字体
    font = xlwt.Font()
    font.name = '微软雅黑'
    font.height = 20 * 12
    style3.font = font
    # 边框
    border = xlwt.Borders()
    border.top = 1
    border.bottom = 1
    border.left = 1
    border.right = 1
    style3.borders = border
    # 对齐方式
    alignment1 = xlwt.Alignment()
    alignment1.vert = 0x01
    alignment1.horz = 0x02
    style3.alignment = alignment1
    # 设置背景颜色
    pattern = xlwt.Pattern()
    pattern.pattern = xlwt.Pattern.SOLID_PATTERN
    pattern.pattern_fore_colour = 3
    style3.pattern = pattern

    # 设置格式4
    style4 = xlwt.XFStyle()
    # 字体
    font = xlwt.Font()
    font.name = '微软雅黑'
    font.height = 20 * 12
    style4.font = font
    # 边框
    border = xlwt.Borders()
    border.top = 1
    border.bottom = 1
    border.left = 1
    border.right = 1
    style4.borders = border
    # 对齐方式
    alignment1 = xlwt.Alignment()
    alignment1.vert = 0x01
    alignment1.horz = 0x02
    style4.alignment = alignment1
    # 设置背景颜色
    pattern = xlwt.Pattern()
    pattern.pattern = xlwt.Pattern.SOLID_PATTERN
    pattern.pattern_fore_colour = 2
    style4.pattern = pattern

    num_pass = 0
    num_fail = 0
    lie_list = []
    module_names = type_name_data.keys()
    for i in module_names:
        nm = len(type_name_data.get(i))
        lie_list.append(nm)
    n = 9
    h = 0
    # 将所有发现的问题写道一起
    result_contents_list = []
    for i, y in type_name_data.items():
        """
        合并单元格使用ws.merge_cells()方法，参数介绍：
        start_row: 开始合并的行。
        start_column: 开始合并的行。
        end_row: 结束合并的列。
        end_column: 结束合并的列。这四个值所在的行/列都会被合并（闭区间）。"""
        # sheet.write(n, 0, i, style2)
        if i == '超级省电（新UX框架）':
            i = '超级省电'
        elif i == '积分系统（卡卡乐园）':
            i = '卡卡乐园'
        elif i == '超级省电（新UX框架）':
            i = '超级省电'
        elif i == '设置-新':
            i = '设置'
        sheet.write_merge(n, n + lie_list[h] - 1, 0, 0, i, style2)
        for x in y:
            # print(x)
            # sheet.write(n, 1, x[0], style)
            sheet.write_merge(n, n, 1, 2, x[0], style)
            # sheet.write(n, 3, x[1], style)
            sheet.write_merge(n, n, 3, 4, x[1], style)
            sheet.write(n, 5, x[2], style2)
            # sheet.write(n, 6, 'Passed', style)
            style_cl = style
            if x[4] == 'Passed':
                num_pass += 1
                style_cl = style3
            elif x[4] == 'Failed':
                num_fail += 1
                style_cl = style4
            sheet.write_merge(n, n, 6, 7, x[4], style_cl)
            if len(x)==6:
                result_contents = str(x[5]).strip()
                if result_contents != '' and result_contents is not None:
                    result_contents_list.append(result_contents + '\n')
                    result_contents = re.findall(r'DTS\d+', str(result_contents))
                sheet.write(n, 8, result_contents, style)
            n += 1
        h += 1
    results = f'Tiger-L10应用特性基本功能用例 执行{num_fail + num_pass}条，Pass:{num_pass}条，Fail:{num_fail}条,通过率{num_pass / (num_fail + num_pass):.2%}。'
    print(results)
    sheet.write(9, 8, results, style)
    if result_contents_list:
        result_contents_list = list(set(result_contents_list))
        print(result_contents_list)
        sheet.write(10, 8, result_contents_list, style)
    # print(results)
    wb.save(csv_file)


if __name__ == '__main__':
    """
    读出执行成功的用例，在全量用例中找出对应的模块和等级，生成执行报告
    """
    # rend_excel()
    # compare_excel()
    write_excel()
    # write_excel_module_classify()
