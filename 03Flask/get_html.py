from lxml import etree
import csv
import os
import openpyxl


class FindCase:

    def __init__(self, file_name):
        self.file_path = file_name
        self.file_save_name = '01本地挂测试报告.xlsx'

    def get_html_date(self):

        html = etree.parse(self.file_path, etree.HTMLParser())
        divs = html.xpath('//table[@class="tests_by_suite"]//tr[@class="test_row"]')
        results_pass = []
        results_file = []
        case_pass = 0
        case_faile = 0
        case_Investigated = 0
        for i in divs:
            case_id = i.xpath('./td[@class="col_name"]/a/text()')[0]
            case_name = i.xpath('./td[@class="col_doc"]/text()')
            if case_name:
                case_name = case_name[0]
            else:
                case_name = '0000'
            result = i.xpath('./td[@class="col_status Passed"]/text()')
            # result = i.xpath("//td[contains(@class 'col_status')]/text()")
            if result == []:
                result = i.xpath('./td[@class="col_status Investigated"]/text()')
            if result == []:
                result = i.xpath('./td[@class="col_status Failed"]/text()')
            result = result[0].strip()
            if result == 'Passed':
                case_pass += 1
            elif result == '':
                case_faile += 1
            elif result == 'Investigated':
                case_Investigated += 1
            # results.append((case_id, case_name, result))
            if result == 'Passed':
                results_pass.append((case_id, case_name, result))
            else:
                results_file.append((case_id, case_name, result))
        # print(results_pass)
        # print(results_file)
        print('脚本执行已完成完,一共{}条用例,成功用例{},失败用例{}'.format(len(results_file + results_pass), len(results_pass),
                                                      len(results_file)))
        print('{:.2%}'.format(case_pass / (len(results_file + results_pass))))
        return results_pass, results_file

    def write_date(self):
        # 删除文件
        if os.path.exists(self.file_save_name):
            os.remove(self.file_save_name)
        results_pass, results_file = self.get_html_date()
        # print(results_pass)
        # print(results_file)
        xls = openpyxl.Workbook()
        sheet = xls.create_sheet('执行成功用例', 0)
        sheet.cell(1, 1, '用例编号')
        sheet.cell(1, 2, '用例名称')
        sheet.cell(1, 3, '执行结果')
        num = 2
        for i in results_pass:
            sheet.cell(num, 1, i[0])
            sheet.cell(num, 2, i[1])
            sheet.cell(num, 3, i[2])
            num += 1

        sheet2 = xls.create_sheet('执行失败用例', 1)
        sheet2.cell(1, 1, '用例编号')
        sheet2.cell(1, 2, '用例名称')
        sheet2.cell(1, 3, '执行结果')
        # print(results_file)
        num2 = 2
        for i in results_file:
            sheet2.cell(num2, 1, i[0])
            sheet2.cell(num2, 2, i[1])
            sheet2.cell(num2, 3, i[2])
            num2 += 1
        xls.save(self.file_save_name)


if __name__ == '__main__':
    file = r'file:///D:/Codephone/BasicPlatform/report/task_2024_11_28_10_17_44/report.html'
    xl = FindCase(file)
    # xl.get_html_date()
    xl.write_date()

"""
from DrissionPage import ChromiumPage
# 定位元素、触发事件
page = ChromiumPage()
page.get('file:///D:/gitfile/BasicPlatform/BasicPlatform/report/task_2024_05_20_20_26_42/report.html')
ele2 = page.ele('xpath:.//table[@class="tests_by_suite"]//tr[@class="test_row"]')
ele3 = page('x:.//table[@class="tests_by_suite"]//tr[@class="test_row"]')
print(ele2)
print(ele3)
# items = page('x:.//table[@class="tests_by_suite"]').eles('x://tr[@class="test_row"]')
items = page.eles('x://tr[@class="test_row"]')
for item in items:
    # print("发件人："+item.ele('t:a').text, "主题："+item.ele('@class:da0').text)
    print(item.ele('x:/td[@class="col_name"]/a').text)
    # print("发件人：" + item.ele('t:a').text)
    print(item.ele('x:/td[@class="col_doc"]').text)
    # print(item.text)


"""
